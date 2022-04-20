#%%
import os
from pyteomics import mzml
import logging
import spectra_plot
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from huvec_analysis import huvec_config

def filter_mzml(mzml_filepath, novel_scan_numbers):
    filtered_mzml = {}
    for mzml_spectra in mzml.MzML(mzml_filepath):
        scan_number = int(mzml_spectra['id'].split('scan=')[1])
        if scan_number in novel_scan_numbers:
            filtered_mzml[scan_number] = mzml_spectra
    return filtered_mzml

def extract_mzml_data(spectra):
    spectra_dict = {}
    try:
        spectra_dict['precursor charge state'] = int(spectra['precursorList']['precursor'][0]['selectedIonList']['selectedIon'][0]['charge state'])
        spectra_dict['precursor m/z'] = float(spectra['precursorList']['precursor'][0]['selectedIonList']['selectedIon'][0]['selected ion m/z'])
    except:
        logging.warn(f"Precuror information not found")
    spectra_dict['lowest observed m/z'] = spectra['lowest observed m/z']
    spectra_dict['highest observed m/z'] = spectra['highest observed m/z']
    spectra_dict['m/z array'] = spectra['m/z array']
    spectra_dict['intensity array'] = spectra['intensity array']
    spectra_dict['base peak m/z'] = spectra['base peak m/z']
    spectra_dict['minimum intensity'] = min(spectra['intensity array'])
    return spectra_dict

def separate_matching_ions(ions, type_convert=float):
    ion_type_list = ions.split(';')
    matched_ions = {}
    for ion_type in ion_type_list:
        ion_map = ion_type.strip('][').split(', ')
        
        for ion in ion_map:
            tmp = ion.split(':')
            matched_ions[tmp[0]] = type_convert(tmp[1])
    return matched_ions

def combine_matches(row):
    combined = {}
    combined = {key:{'m/z':mzml, 'intensity':row['intensity_match'][key]} for key, mzml in row['mz_match'].items()}
    # for key in mzml.keys():
    #     print(key)
    #     print(mzml)
    #     return mzml
    #     combined[key]['m/z'] = mzml[key]
    #     combined[key]['intensity'] = intensity[key]
    return combined
def process_metamorpheus_psm(psm_filepath, novel_scans):
    psm = pd.read_table(psm_filepath)
    filtered_psms = []
    for scan in novel_scans:
        tmp = psm
        for column, value in scan.items():
            tmp = tmp[tmp[column] == value]
        if len(tmp) > 0:
            filtered_psms.append(tmp)
    if len(filtered_psms) > 0:
        psm = pd.concat(filtered_psms)
    else:
        raise IndexError("No valid PSMs found")
    # psm = psm[psm['Scan Number'].isin(novel_peptides)]
    psm['mz_match'] = psm['Matched Ion Mass-To-Charge Ratios'].apply(separate_matching_ions)
    psm['intensity_match'] = psm['Matched Ion Intensities'].apply(separate_matching_ions)
    psm['matched_ions'] = psm.apply(combine_matches, axis = 1)
    return psm 


def plot_mirror_into_exel(psm, mzml_directory):
    if not os.path.exists('plot'):
        os.mkdir('plot')
    if not os.path.exists('plot/individual_mirror_plots'):
        os.mkdir('plot/individual_mirror_plots')


    filename_column = 'A'
    scan_number_column = 'B'
    peptide_base_column = 'C'
    peptide_full_column = 'D'
    accession_column = 'E'
    gene_column = 'F'
    score_column = 'G'
    # qvalue_column = ''
    precursor_column = 'H'
    psm_count_column = 'I'
    mass_diff_columns = 'J'
    mirrorplot_column='K'
    
    wb = Workbook()
    ws = wb.active

    ws.append([
        "Filename", 
        "Scan Number", 
        "Base Sequence",
        "Full Sequence", 
        "Accession",
        "Gene", 
        "Score", 
        # "Q-Value", 
        "Precursor Charge", 
        "PSM Count (unambiguous, <0.01 q-value)", 
        "Mass Diff (ppm)", 
        "Mirror Plot"])
    ws.row_dimensions[2].height=375
    ws.column_dimensions[mirrorplot_column].width = 80
    excel_row_number = 2
    for filename, group in psm.groupby('File Name'):
        scans = list(group['Scan Number'])

        mzml_filepath=os.path.join(mzml_directory, filename) + '.mzML'
        novel_mzml = filter_mzml(mzml_filepath, scans)
        for index, row in group.iterrows():
            scan = row['Scan Number']
            # print(scan)
            # mzml_data = extract_mzml_data(novel_mzml[scan])
            # mirror = spectra_plot.SpectraPlot(
            #     mass_charge_ratios=mzml_data['m/z array'],
            #     intensities = mzml_data['intensity array'])
            # mirror.labelled_spectra = row['matched_ions']
            # mirror.filter()
            # ax = mirror.mirror_plot(label=True,adjust_annotation = False) #TODO fix plot size fig.set_size_inches(18.5, 10.5)
            # plt.savefig(f'plot/individual_mirror_plots/mass-spec_{scan}.png')
            # img = Image(f'plot/individual_mirror_plots/mass-spec_{scan}.png')
            # ws.add_image(img, f'{mirrorplot_column}{excel_row_number}')
            ws.row_dimensions[excel_row_number].height=375

            ws[f'{filename_column}{excel_row_number}']=row['File Name']
            ws[f'{scan_number_column}{excel_row_number}']=row['Scan Number']
            ws[f'{peptide_base_column}{excel_row_number}']=row['Base Sequence']
            ws[f'{peptide_full_column}{excel_row_number}']=row['Full Sequence']
            ws[f'{accession_column}{excel_row_number}']=row['Protein Accession']
            ws[f'{gene_column}{excel_row_number}']=row['Gene Name']
            ws[f'{score_column}{excel_row_number}']=row['Score']
            ws[f'{precursor_column}{excel_row_number}']=row['Precursor Charge']
            ws[f'{psm_count_column}{excel_row_number}']=row['PSM Count (unambiguous, <0.01 q-value)']
            ws[f'{mass_diff_columns}{excel_row_number}']=row['Mass Diff (ppm)']
            excel_row_number = excel_row_number + 1
    wb.save(f'stats/huvec_novel_peptides.xlsx')

psm_filepath = '../00_pre_analysis/metamorpheus_table/AllPSMs.PacBioHybrid.tsv'
novel_filepath = 'stats/pacbio_hybrid_novel_peptides.tsv'


novel_scans = pd.read_table(novel_filepath)

novel_scans_list = []
for index, row in novel_scans.iterrows():
    novel_scans_list.append({'File Name' : row['File Name'], 'Scan Number': row['Scan Number']})
psm = process_metamorpheus_psm(psm_filepath, novel_scans_list)

plot_mirror_into_exel( psm, huvec_config.MZML_DIRECTORY)

